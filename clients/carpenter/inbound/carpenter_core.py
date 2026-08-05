"""
carpenter.py — Extraction engine for Carpenter Technology shipment documents.

Parses:
  1. Order List Excel (Receipt_Report.xlsx)
  2. Arrival Notice PDF (ACL format)
  3. N × Packing List PDFs (Carpenter Technology format)

Merges into a single Excel output with updated columns.

Changes (latest):
  - Added "Customer" column (always "CARPENTER (CPTR)")
  - Cost Center updated to "O-VMROOUDENDIJK"
  - Added "Operation" column (always "Discharging ex-container")
  - "Confirmation Purpose" → "Rapportage Klant" (corrected spelling)
  - "Product" → "Colli" (was "Collis")
  - Added "PALLET" and "PIECES" columns
  - Weight extracted from Weight column (LB/KG pair): takes KG value (after "/"),
    truncated to 2 decimal digits (no rounding)
  - PALLET / PIECES: parsed from package summary block label (e.g. "1 BOX", "2 PALLET")
    → if count == 0 → 0 ; if count > 0 → 1 (the word BOX/PALLET/etc. is ignored)

  BUGFIX (this version):
  - Weight-column extraction was bleeding into the adjacent Length column because
    the x0 upper bound (700) overlapped both columns. Weight-column data sits at
    roughly x0 619-680; Length-column data starts at x0 ~695. The bound has been
    tightened to 687 so Length text is never appended to weight_raw. This fixes
    garbled/incorrect Gross/Net Weight (KG) values such as "72833.17" appearing
    for packages that should be a few hundred kg.
  - The per-row anchor pattern for the "Prod Order / MF Order" column only matched
    pure digit strings (^\\d{7,8}$). Some shipment batches (e.g. Reading Operations
    plant deliveries) use letter-prefixed codes there instead (e.g. "L20165",
    "W94727"), which never matched — causing entire packing list PDFs to yield
    zero extracted rows. The pattern now also accepts an optional 1-2 letter
    prefix: ^[A-Z]{0,2}\\d{5,8}$.

  BUGFIX (v2 — seal & instruction):
  - Seal-number regex in parse_arrival_notice was ^UL\\d{7}$, which missed
    hyphenated seals like "UL-6101303". Now uses ^UL-?\\d{7}$ (optional hyphen).
  - Seal search was per-page (25-word window), so containers at a page break
    (e.g. GCNU4859065 on page 2, seal on page 3) were missed. Now combines
    words from all pages into one sorted list and scans forward until the next
    container — handles any cross-page split.
  - Warehouse Instruction cleaning: "SHIP ASAP - WEEK 25" → "SHIP ASAP" (strip
    week/date suffix); "NO INS" → "VMR STOCK"; "KEEP VM" / "KEPT VM" → "VMR STOCK".
"""

import re
import os
import math
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from clients.carpenter.inbound.scanned_doc import extract_scanned_pages, link_scanned_to_deliveries
from clients.carpenter.inbound.arrival_notice import parse_arrival_notice_llm


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def _is_valid_doc_no(s: str) -> bool:
    """
    Return True if s looks like a valid Carpenter document number:
      - 7 or 8 digits
      - starts with '100' or '1000'

    Examples:
        '1007417'   → True
        '10038897'  → True
        '80588261'  → False  (delivery number, not doc no)
        'L20165'    → False
    """
    return bool(re.match(r"^\d{7,8}$", s)) and s.startswith(("100", "1000"))


def _truncate2(value_str: str) -> str:
    """
    Extract the KG value from a LB/KG pair string like '955.000/433.180'
    or a plain number string, then truncate to 2 decimal places (NO rounding).

    Examples:
        '955.000/433.180'  → '433.18'
        '510'              → '510'
        '509.837'          → '509.83'
        ''                 → ''
    """
    if not value_str:
        return ""

    raw = str(value_str).strip()

    # If the string contains "/" it's a LB/KG pair — take the part after "/"
    if "/" in raw:
        raw = raw.split("/")[-1].strip()

    # Remove commas used as thousands separators
    raw = raw.replace(",", "")

    try:
        f = float(raw)
    except ValueError:
        return raw  # return as-is if unparseable

    # Truncate (not round) to 2 decimal places
    truncated = math.floor(f * 100) / 100

    # Format: drop trailing zeros after decimal but keep at most 2 places
    if truncated == int(truncated):
        return str(int(truncated))
    else:
        # Render with 2 decimals, then strip trailing zero if only 1 decimal needed
        s = f"{truncated:.2f}"
        return s.rstrip("0").rstrip(".")  # e.g. "433.10" → "433.1", "433.18" stays


def _parse_pkg_count(label: str) -> int:
    """
    Parse the numeric prefix from a package label like '1 BOX', '2 PALLET', '0 PIECES'.
    Returns the integer count. Returns 0 if no number found.

    Examples:
        '1 BOX'     → 1
        '2 PALLET'  → 2
        '0'         → 0
        'BOX'       → 0
    """
    if not label:
        return 0
    m = re.match(r"(\d+)", str(label).strip())
    return int(m.group(1)) if m else 0


def _count_to_binary(count: int) -> int:
    """If count > 0 → 1, else → 0."""
    return 1 if count > 0 else 0


def _clean_instruction(raw) -> str:
    """
    Clean warehouse instruction value:
      - 'SHIP ASAP - WEEK 25'              → 'SHIP ASAP'  (strip any suffix after -)
      - 'SHIP ASAP -  MID-JUNE, CWD ...'   → 'SHIP ASAP'
      - 'SHIP ASAP - JUNE: CWD ...'        → 'SHIP ASAP'
      - 'NO INS'                           → 'VMR STOCK'
      - 'KEEP VM' / 'KEPT VM'              → 'VMR STOCK'
      - NaN / empty                        → ''
    """
    if pd.isna(raw) or raw is None:
        return ""
    val = str(raw).strip()
    if not val:
        return ""

    upper = val.upper()

    if upper in ("NO INS", "KEEP VM", "KEPT VM"):
        return "VMR STOCK"

    # Strip any trailing suffix after "SHIP ASAP - ..."
    m = re.match(r"^(SHIP\s+ASAP)\s*[-–].*", upper)
    if m:
        return "SHIP ASAP"

    return val  # return original casing for other values


# ═══════════════════════════════════════════════════════════════
# 1. ORDER LIST EXCEL PARSER
# ═══════════════════════════════════════════════════════════════

def parse_order_list(xlsx_path: str) -> pd.DataFrame:
    """Read Order List Excel and return cleaned DataFrame."""
    df = pd.read_excel(xlsx_path)

    col_map = {}
    for col in df.columns:
        cl = col.strip().lower()
        if "receipt" in cl and "p.o" in cl:
            col_map[col] = "receipt_po"
        elif "order" in cl and "no" in cl:
            col_map[col] = "order_no"
        elif "units" in cl and "expected" in cl:
            col_map[col] = "units_expected"
        elif "gross" in cl and "weight" in cl:
            col_map[col] = "total_gross_kg"
        elif "box" in cl and "range" in cl:
            col_map[col] = "end_customer"
        elif "instruction" in cl:
            col_map[col] = "warehouse_instruction"
        elif "oi" in cl and "76" in cl:
            col_map[col] = "ref_no"
        elif "hbl" in cl:
            col_map[col] = "hbl"
        elif "conveyance" in cl:
            col_map[col] = "container_no"
        elif "vessel" in cl:
            col_map[col] = "vessel"
        elif "receipt" in cl and "no" in cl:
            col_map[col] = "receipt_no"

    df = df.rename(columns=col_map)
    df["receipt_po"] = df["receipt_po"].astype(str).str.strip()
    df["container_no"] = df["container_no"].astype(str).str.strip()
    df["ref_no"] = df["ref_no"].astype(str).str.strip()

    # ── Split multi-delivery receipt_po cells ──
    # Some rows have "80588261 80588262" (space-separated) in Receipt P.O.
    # Split into one row per delivery so that merge_data can match each
    # delivery_no individually.  The first delivery keeps units_expected
    # as-is; additional deliveries get units_expected = 0 (they are
    # duplicate/companion HUs that should not count toward validation).
    expanded_rows = []
    for _, row in df.iterrows():
        po_val = str(row.get("receipt_po", "")).strip()
        parts = po_val.split()
        if len(parts) > 1:
            for idx, part in enumerate(parts):
                new_row = row.copy()
                new_row["receipt_po"] = part
                # First delivery keeps units_expected; the rest are duplicates
                if idx > 0:
                    new_row["units_expected"] = 0
                expanded_rows.append(new_row)
        else:
            expanded_rows.append(row)

    if expanded_rows:
        df = pd.DataFrame(expanded_rows).reset_index(drop=True)

    return df


# ═══════════════════════════════════════════════════════════════
# 2. ARRIVAL NOTICE PDF PARSER
# ═══════════════════════════════════════════════════════════════

def parse_arrival_notice(pdf_path: str) -> dict:
    """
    Parse ACL Arrival Notice PDF.
    Returns dict: container_no → {eta, etd, seal_no, weight_kgs}

    FIX v2:
      - Seal pattern now accepts optional hyphen: ^UL-?\\d{7}$
      - Words from ALL pages are combined into a single sorted list so that
        containers whose seal falls on the next page (cross-page split) are
        handled correctly.
      - Seal search scans forward until the next container number (instead of
        a fixed 25-word window), matching only words at x0 < 50 (the left
        column where seals appear).
    """
    pdf = pdfplumber.open(pdf_path)
    container_pattern = re.compile(r"^[A-Z]{4}\d{7}$")
    seal_pattern = re.compile(r"^UL-?\d{7}$")          # ← FIX: optional hyphen

    # ── ETA / ETD from page 1 ──
    page1_words = pdf.pages[0].extract_words()
    eta, etd = "", ""
    for i, w in enumerate(page1_words):
        if w["text"] == "ETA:" and i + 2 < len(page1_words):
            parts = [page1_words[j]["text"] for j in range(i + 1, min(i + 3, len(page1_words)))]
            eta = " ".join(parts)
        if w["text"] == "ETD:" and i + 2 < len(page1_words):
            parts = [page1_words[j]["text"] for j in range(i + 1, min(i + 3, len(page1_words)))]
            etd = " ".join(parts)

    # ── Combine words from ALL pages into one globally-sorted list ──
    # This ensures cross-page container/seal pairs are found.
    all_words = []
    for pi, page in enumerate(pdf.pages):
        for w in page.extract_words():
            all_words.append({
                **w,
                "global_top": pi * 10000 + w["top"],
            })
    all_words.sort(key=lambda w: (w["global_top"], w["x0"]))

    # ── Extract containers + seals ──
    containers = {}
    for i, w in enumerate(all_words):
        if container_pattern.match(w["text"]) and w["x0"] < 120:
            ctr = w["text"]
            if ctr in containers:
                continue

            # Scan forward for the seal. Stop at the next container number.
            # Only match seal words in the left column (x0 < 50) where seals
            # always appear, so page-header noise is skipped.
            seal = ""
            for j in range(i + 1, min(i + 150, len(all_words))):
                candidate = all_words[j]
                # Stop if we hit another container number in the left column
                if container_pattern.match(candidate["text"]) and candidate["x0"] < 120:
                    break
                if seal_pattern.match(candidate["text"]) and candidate["x0"] < 50:
                    seal = candidate["text"]
                    break

            # Weight in KGS on the same line
            weight_kgs = ""
            for j in range(i + 1, min(i + 6, len(all_words))):
                if (all_words[j]["text"] == "KGS"
                        and abs(all_words[j]["global_top"] - w["global_top"]) < 3):
                    weight_kgs = all_words[j - 1]["text"]
                    break

            containers[ctr] = {
                "container_no": ctr,
                "eta": eta,
                "etd": etd,
                "seal_no": seal,
                "weight_kgs": weight_kgs,
            }

    pdf.close()
    return containers


# ═══════════════════════════════════════════════════════════════
# 3. PACKING LIST PDF PARSER
# ═══════════════════════════════════════════════════════════════

def _derive_container_from_filename(filename: str) -> str:
    """
    Extract container number from filename like 'ACLU9779950.pdf'.
    Returns empty string if no valid container pattern found (e.g.
    'Packing_lists.pdf') — the linker will resolve container_no
    from the order list instead.
    """
    basename = os.path.splitext(os.path.basename(filename))[0]
    match = re.search(r"[A-Z]{4}\d{7}", basename)
    return match.group(0) if match else ""


def _detect_category(words: list) -> str:
    """
    Check SHIP FROM address block on first data page.
    If the address contains 'Carpenter' → 'carpenter', else blank.

    The SHIP FROM section is on the right side of the packing list header.
    We find the "SHIP" + "FROM" label dynamically, then read the address
    lines below it and check for "CARPENTER".
    """
    # ── Strategy 1: Find "SHIP" + "FROM" label dynamically ──
    ship_from_top = None
    ship_from_x0 = None
    for w in words:
        if w["text"].upper() == "SHIP" and w["top"] < 300:
            for w2 in words:
                if (w2["text"].upper() == "FROM"
                        and abs(w2["top"] - w["top"]) < 5
                        and w2["x0"] > w["x0"]
                        and w2["x0"] - w["x0"] < 80):
                    ship_from_top = w["top"]
                    ship_from_x0 = w["x0"]
                    break
        if ship_from_top is not None:
            break

    if ship_from_top is not None:
        # Collect address words below "SHIP FROM"
        address_words = []
        for w in words:
            if (ship_from_top + 3 < w["top"] < ship_from_top + 80
                    and w["x0"] >= ship_from_x0 - 10):
                address_words.append(w)
        address_words.sort(key=lambda w: (w["top"], w["x0"]))
        address_text = " ".join(w["text"] for w in address_words).upper()
        if "LATROBE" in address_text:
            return "latrobe"
        if "CARPENTER" in address_text:
            return "carpenter"

    # ── Strategy 2 (fallback): scan all header text for CARPENTER ──
    header_text = " ".join(
        w["text"] for w in words if w["top"] < 250
    ).upper()
    if "LATROBE" in header_text:
        return "latrobe"
    if "CARPENTER" in header_text:
        return "carpenter"

    return ""


def parse_packing_list(pdf_path: str) -> tuple:
    """
    Parse a Carpenter Technology Packing List PDF.

    Returns:
        (hu_records, validation_info)

        hu_records: list of dicts, one per Handling Unit
        validation_info: {delivery_no: hu_count} for validation

    Weight extraction:
        The Weight column in line-item rows contains a LB/KG pair like '955.000/433.180'.
        We extract the KG part (after '/') and truncate to 2 decimal places.

        NOTE: the Weight column (x0 ~619-680) sits immediately to the left of the
        Length column (x0 starts ~695). The scan window is capped at x0 < 687 so
        Length-column text is never appended onto the weight string.

    Package label parsing (PALLET / PIECES):
        The package summary block contains a label like '1 BOX', '2 PALLET', etc.
        We parse the numeric prefix:
          - count == 0 → PALLET=0, PIECES=0
          - count > 0 and keyword is PALLET → PALLET=1, PIECES=0
          - count > 0 and keyword is anything else (BOX, etc.) → PALLET=0, PIECES=1
        If no keyword matches PALLET, treat as PIECES.
    """
    pdf = pdfplumber.open(pdf_path)
    container = _derive_container_from_filename(pdf_path)
    delivery_pattern = re.compile(r"^80\d{6}$")

    category = ""
    for page in pdf.pages:
        page_words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))
        if page_words:
            category = _detect_category(page_words)
            if category:
                break

    # ── Pass 1: Extract line item rows (batch-level) ──
    current_delivery = None
    line_items = []
    delivery_order_map = {}

    for page_idx, page in enumerate(pdf.pages):
        words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))

        for w in words:
            if delivery_pattern.match(w["text"]) and 95 < w["x0"] < 150 and w["top"] < 350:
                current_delivery = w["text"]
                break

        if not current_delivery:
            continue

        # "Order No" anchor column (actually the Prod Order / MF Order field).
        # Some shipment batches (e.g. Reading Operations plant) use letter-prefixed
        # codes here (e.g. "L20165", "W94727") instead of plain digits, so the
        # pattern accepts an optional 1-2 letter prefix.
        order_no_pattern = re.compile(r"^[A-Z]{0,2}\d{5,8}(-\d{1,2})?$")

        for w in words:
            if 510 < w["x0"] < 560 and order_no_pattern.match(w["text"]):
                order_no = w["text"]
                order_top = w["top"]
                hu = None
                material = None
                batch = None
                weight_raw = None
                weight_top = None

                for w2 in words:
                    if abs(w2["top"] - order_top) < 3:
                        # Handling Unit column (x0 ≈ 569)
                        if 565 < w2["x0"] < 625 and w2["text"] not in ("Handling", "Unit"):
                            if len(w2["text"]) >= 5:
                                hu = w2["text"]
                        # Batch column (x0 ≈ 281)
                        if 275 < w2["x0"] < 340 and len(w2["text"]) > 5:
                            batch = w2["text"]

                # ─────────────────────────────────────────────────────────────

                weight_raw = ""

                # Find the first part containing "/".
                # Upper bound tightened to 687 (was 700) so we stop before the
                # Length column (which starts at x0 ~695) and never mix the two.
                for w2 in words:
                    if (
                        "/" in w2["text"]
                        and w2["text"] != "LB/KG"
                        and 590 < w2["x0"] < 687
                        and abs(w2["top"] - order_top) < 12
                    ):
                        weight_raw = w2["text"]
                        weight_top = w2["top"]

                        # Same-line continuation
                        same_line = sorted(
                            [
                                x for x in words
                                if abs(x["top"] - weight_top) < 3
                                and x["x0"] > w2["x0"]
                                and x["x0"] < 687
                            ],
                            key=lambda x: x["x0"]
                        )

                        for x in same_line:
                            if x["text"] != "LB/KG":
                                weight_raw += x["text"]

                        # Next-line continuation
                        next_line = sorted(
                            [
                                x for x in words
                                if 2 < (x["top"] - weight_top) < 18
                                and 590 < x["x0"] < 687
                            ],
                            key=lambda x: x["x0"]
                        )

                        for x in next_line:
                            if x["text"] != "LB/KG":
                                weight_raw += x["text"]

                        # Extract only LB/KG pair
                        m = re.search(
                            r'(\d[\d,]*\.\d+\s*/\s*\d[\d,]*\.\d+)',
                            weight_raw.replace(" ", "")
                        )

                        if m:
                            weight_raw = m.group(1)

                        break

                # Document No (Material column, x0 ≈ 101)
                # Collect ALL valid doc-number candidates in the vertical window
                # around this order row.  The Material # column sometimes prints
                # two stacked numbers, e.g.:
                #   1007417     ← top ≈ T
                #   10038897    ← top ≈ T + 12
                # We combine them as "1007417/10038897".
                doc_candidates = []
                for w2 in words:
                    if (
                        98 < w2["x0"] < 200
                        and w["top"] - 35 < w2["top"] < w["top"] + 20
                        and _is_valid_doc_no(w2["text"])
                    ):
                        doc_candidates.append(w2)

                # Sort top-to-bottom so the first line always comes first
                doc_candidates.sort(key=lambda x: x["top"])

                if len(doc_candidates) >= 2:
                    # Combine first two valid numbers (extra ones are ignored)
                    material = f"{doc_candidates[0]['text']}/{doc_candidates[1]['text']}"
                elif len(doc_candidates) == 1:
                    material = doc_candidates[0]["text"]
                # else: material stays None (no valid doc no found)

                if hu and order_no:
                    line_items.append({
                        "delivery_no": current_delivery,
                        "order_no": order_no,
                        "handling_unit": hu,
                        "document_no": material or "",
                        "batch": batch or "",
                        "weight_raw": weight_raw or "",
                    })

                    if current_delivery not in delivery_order_map:
                        delivery_order_map[current_delivery] = set()
                    delivery_order_map[current_delivery].add(order_no)

    # ── Pass 2: Extract package summary blocks (HU-level) ──
    current_delivery_pkg = None
    package_summaries = []

    for page_idx, page in enumerate(pdf.pages):
        words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))

        for w in words:
            if delivery_pattern.match(w["text"]) and 95 < w["x0"] < 150 and w["top"] < 350:
                current_delivery_pkg = w["text"]
                break

        for w in words:
            if w["text"] == "Handling" and w["x0"] < 60:
                handling_top = w["top"]

                # Package type label (e.g. "BOX", "PALLET") on same line at x0 > 120
                pkg_type_words = []
                for w2 in words:
                    if abs(w2["top"] - handling_top) < 3 and 120 < w2["x0"] < 230:
                        if w2["text"] not in ("Handling", "Package", "Type"):
                            pkg_type_words.append(w2["text"])
                pkg_type = " ".join(pkg_type_words)

                # HU number slightly below "Handling" at x0 ≈ 69
                hu_num = None
                hu_top = None
                for w2 in words:
                    if handling_top < w2["top"] < handling_top + 12 and 60 < w2["x0"] < 120:
                        if len(w2["text"]) >= 7 and w2["text"] != "Unit":
                            hu_num = w2["text"]
                            hu_top = w2["top"]
                            break

                if not hu_num or not hu_top:
                    continue

                # Gross/Net weight are NOT read from the summary block anymore.
                # Both output columns come from the line item's "Weight" column,
                # linked in via hu_lookup[hu]["weight_raw"] below.
                gross_weight_combined = ""
                net_weight_combined = ""

                # ── Package count label parsing ──
                # Strategy 1: Look for "N BOX" / "N PALLET" label below the HU block
                pkg_label = ""
                pkg_label_pattern = re.compile(r"^\d+$")
                for w2 in words:
                    if hu_top + 30 < w2["top"] < hu_top + 120 and 50 < w2["x0"] < 200:
                        if pkg_label_pattern.match(w2["text"]):
                            for w3 in words:
                                if abs(w3["top"] - w2["top"]) < 3 and w3["x0"] > w2["x0"]:
                                    pkg_label = f"{w2['text']} {w3['text']}"
                                    break
                            if not pkg_label:
                                pkg_label = w2["text"]
                            break

                pkg_count = _parse_pkg_count(pkg_label)

                # An explicit "0 <label>" (e.g. "0 SKID/LID") is the PDF telling us
                # there is no pallet/skid for this HU — Strategy 2 must NOT override
                # that real zero with an unrelated "Contains N Pieces" count.
                has_explicit_zero = (
                    pkg_label
                    and re.match(r"^0\b", pkg_label.strip())
                )

                # Strategy 2: Look for "Contains X Pieces" near the HU block
                if pkg_count == 0 and not has_explicit_zero:
                    for w2 in words:
                        if (w2["text"] == "Contains"
                                and handling_top - 5 < w2["top"] < hu_top + 30):
                            for w3 in words:
                                if (abs(w3["top"] - w2["top"]) < 3
                                        and w3["x0"] > w2["x0"]
                                        and re.match(r"^\d+$", w3["text"])):
                                    pkg_count = int(w3["text"])
                                    pkg_label = f"{w3['text']} Pieces"
                                    break
                            break

                if pkg_count == 0 and hu_num and not has_explicit_zero:
                    pkg_count = 1

                # Any non-zero count → PALLET=1, PIECES=1
                if pkg_count > 0:
                    pallet_val = 1
                    pieces_val = 1
                else:
                    pallet_val = 0
                    pieces_val = 0

                package_summaries.append({
                    "delivery_no": current_delivery_pkg,
                    "handling_unit": hu_num,
                    "package_type": pkg_type or "",
                    "gross_weight_raw": gross_weight_combined,
                    "net_weight_raw": net_weight_combined,
                    "container_no": container,
                    "category": category,
                    "pallet": pallet_val,
                    "pieces": pieces_val,
                    "pkg_label": pkg_label,
                })

    # ── Linking: connect HU → Order No + Document No + weight ──
    hu_lookup = {}
    for li in line_items:
        hu = li["handling_unit"]
        if hu not in hu_lookup:
            hu_lookup[hu] = {
                "order_no": li["order_no"],
                "document_no": li["document_no"],
                "weight_raw": li["weight_raw"],
            }
        if li["document_no"] and not hu_lookup[hu]["document_no"]:
            hu_lookup[hu]["document_no"] = li["document_no"]
        if li["weight_raw"] and not hu_lookup[hu]["weight_raw"]:
            hu_lookup[hu]["weight_raw"] = li["weight_raw"]

    delivery_to_order = {}
    for li in line_items:
        if li["delivery_no"] not in delivery_to_order:
            delivery_to_order[li["delivery_no"]] = li["order_no"]

    delivery_to_docs = {}
    for li in line_items:
        if li["document_no"] and li["document_no"].startswith(("100", "1000")):
            delivery_to_docs.setdefault(li["delivery_no"], set()).add(li["document_no"])

    delivery_to_weight = {}
    for li in line_items:
        if li["weight_raw"] and li["delivery_no"] not in delivery_to_weight:
            delivery_to_weight[li["delivery_no"]] = li["weight_raw"]

    for ps in package_summaries:
        hu = ps["handling_unit"]
        if hu in hu_lookup:
            ps["order_no"] = hu_lookup[hu]["order_no"]
            doc = hu_lookup[hu]["document_no"]
            if doc and doc.startswith(("100", "1000")):
                ps["document_no"] = doc
            else:
                docs = delivery_to_docs.get(ps["delivery_no"], set())
                ps["document_no"] = next(iter(docs), "")
            ps["gross_weight_raw"] = hu_lookup[hu]["weight_raw"]
            ps["net_weight_raw"] = hu_lookup[hu]["weight_raw"]
        else:
            ps["order_no"] = delivery_to_order.get(ps["delivery_no"], "")
            docs = delivery_to_docs.get(ps["delivery_no"], set())
            ps["document_no"] = next(iter(docs), "")
            # NEW: give weight a fallback too, instead of leaving it blank
            fallback_weight = delivery_to_weight.get(ps["delivery_no"], "")
            ps["gross_weight_raw"] = fallback_weight
            ps["net_weight_raw"] = fallback_weight

    validation = {}
    for ps in package_summaries:
        d = ps["delivery_no"]
        validation[d] = validation.get(d, 0) + 1

    pdf.close()
    return package_summaries, validation


# ═══════════════════════════════════════════════════════════════
# 4. MERGE + OUTPUT
# ═══════════════════════════════════════════════════════════════

def merge_data(
    df_orders: pd.DataFrame,
    arrival: dict,
    all_hu_records: list,
) -> pd.DataFrame:
    """
    Merge all three sources into the final DataFrame.

    Column order:
      Ref No | ETA Date | Container No | Customer | Cost Center | Seal No |
      Confirmation Purpose | Product | Operation | Order No | Handling Unit |
      Delivery No | Category | End Customer | Customer Status |
      Warehouse Instruction | Document No | Doc Type |
      Gross Weight (KG) | Net Weight (KG) | PALLET | PIECES
    """
    df_hu = pd.DataFrame(all_hu_records)

    if df_hu.empty:
        return pd.DataFrame()

    # Merge with Excel on delivery_no = receipt_po
    # Merge with Excel on delivery_no = receipt_po AND order_no (prevents fan-out
    # when one receipt_po has multiple rows with different order_nos)
    df_orders_norm = df_orders.copy()
    df_orders_norm["order_no_norm"] = (
        df_orders_norm["order_no"]
        .astype(str)
        .str.replace(r"(?i)^wo", "", regex=True)
        .str.strip()
    )
    df_hu["order_no_norm"] = (
        df_hu["order_no"]
        .astype(str)
        .str.replace(r"(?i)^wo", "", regex=True)
        .str.strip()
    )

    df_merged = df_hu.merge(
        df_orders_norm[[
            "receipt_po", "order_no_norm", "units_expected",
            "end_customer", "warehouse_instruction",
            "ref_no", "hbl"
        ]],
        left_on=["delivery_no", "order_no_norm"],
        right_on=["receipt_po", "order_no_norm"],
        how="left",
        suffixes=("_pkg", "_excel"),
    )

    print(f"  [DEBUG] df_hu rows: {len(df_hu)}")
    print(f"  [DEBUG] df_orders_norm sample:\n{df_orders_norm[['receipt_po','order_no','order_no_norm']].to_string()}")
    print(f"  [DEBUG] df_merged rows after merge: {len(df_merged)}")
    print(df_merged[["delivery_no", "order_no_norm", "handling_unit", "receipt_po"]].to_string())

    df_merged["order_no_pkg"] = df_hu["order_no_norm"].values
    df_merged["order_no_excel"] = df_merged["order_no_norm"]

    # ── Fallback for unmatched rows (receipt_po is NaN after dual-key merge) ──
    # These are HUs where the PDF's order_no doesn't match the Excel's order_no
    # (e.g. Excel stores Prod Order but PDF stores MF Order or vice versa).
    # For these, match on delivery_no alone using the FIRST Excel row for that PO
    # to get ref_no, end_customer, warehouse_instruction etc.
    unmatched = df_merged["receipt_po"].isna() & (df_merged.get("source", "") != "scanned")
    if unmatched.any():
        # One row per receipt_po — take first (preserves ref_no, end_customer etc.)
        df_orders_first = (
            df_orders_norm
            .drop_duplicates(subset=["receipt_po"], keep="first")
        )
        # Re-merge only the metadata columns (not order_no_norm — keep pkg value)
        df_meta = df_hu[unmatched][["delivery_no"]].merge(
            df_orders_first[[
                "receipt_po", "units_expected",
                "end_customer", "warehouse_instruction",
                "ref_no", "hbl"
            ]],
            left_on="delivery_no",
            right_on="receipt_po",
            how="left",
        )
        # Fill only the NaN metadata columns — do NOT overwrite order_no_norm
        for col in ["receipt_po", "units_expected", "end_customer",
                    "warehouse_instruction", "ref_no", "hbl"]:
            if col in df_meta.columns and col in df_merged.columns:
                df_merged.loc[unmatched, col] = df_meta[col].values

    # Merge with Arrival Notice on container_no
    df_arrival = pd.DataFrame(arrival.values())
    if not df_arrival.empty:
        df_merged = df_merged.merge(
            df_arrival[["container_no", "eta", "etd", "seal_no"]],
            on="container_no",
            how="left",
        )
    else:
        df_merged["eta"] = ""
        df_merged["etd"] = ""
        df_merged["seal_no"] = ""

    # Use pkg order_no (from PDF), fallback to Excel
    order_pkg = df_merged.get("order_no_pkg", pd.Series(dtype=str))
    order_excel = df_merged.get("order_no_excel", pd.Series(dtype=str))
    df_merged["order_no_final"] = order_pkg.where(order_pkg.notna() & (order_pkg != ""), order_excel)

    # Apply weight truncation (KG, 2 decimal places, no rounding)
    gross_col = df_merged.get("gross_weight_raw", pd.Series(dtype=str)).fillna("")
    net_col = df_merged.get("net_weight_raw", pd.Series(dtype=str)).fillna("")
    gross_kg_series = gross_col.apply(_truncate2)
    net_kg_series = net_col.apply(_truncate2)

    # Build final output
    final = pd.DataFrame()
    final["Ref No"] = df_merged.get("ref_no", "")
    final["ETA Date"] = df_merged.get("eta", "")
    final["Container No"] = df_merged.get("container_no", "")
    final["Customer"] = "CARPENTER (CPTR)"                      # ← NEW (always fixed)
    final["Cost Center"] = "O-VMROOUDENDIJK"                    # ← UPDATED
    final["Seal No"] = df_merged.get("seal_no", "")
    final["Confirmation Purpose"] = "Rapportage Klant"          # ← CORRECTED spelling
    final["Product"] = "Colli"                                  # ← UPDATED (was Collis)
    final["Operation"] = "Discharging ex-container"             # ← NEW (always fixed)
    final["Order No"] = df_merged.get("order_no_final", "")
    final["Handling Unit"] = df_merged.get("handling_unit", "")
    final["Delivery No"] = df_merged.get("delivery_no", "")
    final["Category"] = df_merged.get("category", "")
    final["End Customer"] = df_merged.get("end_customer", "")
    final["Customer Status"] = "BONDED"
    # ← FIX v2: clean instruction values
    final["Warehouse Instruction"] = df_merged.get("warehouse_instruction", "").fillna("").apply(_clean_instruction)
    final["Document No"] = df_merged.get("document_no", "")
    final["Doc Type"] = "IMAJ"
    final["Gross Weight (KG)"] = gross_kg_series               # ← KG truncated to 2dp
    final["Net Weight (KG)"] = net_kg_series                   # ← KG truncated to 2dp
    final["PALLET"] = df_merged.get("pallet", 0)               # ← NEW
    final["PIECES"] = df_merged.get("pieces", 0)               # ← NEW

    return final


def write_output_excel(df: pd.DataFrame, output_path: str):
    """Write final DataFrame to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipment Data"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value) or value == "" or value is None:
                clean_val = ""
            else:
                clean_val = str(value)
                if clean_val.endswith(".0") and clean_val[:-2].isdigit():
                    clean_val = clean_val[:-2]

            cell = ws.cell(row=row_idx, column=col_idx, value=clean_val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(df.columns[col_idx - 1]))
        for row_idx in range(2, len(df) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════
# 5. VALIDATION
# ═══════════════════════════════════════════════════════════════

def validate(df_orders: pd.DataFrame, all_hu_records: list) -> list:
    """
    Run validation checks. Returns list of message strings.

    Duplicate HUs (pallet == 0 AND pieces == 0) are excluded from the
    count because the order list only counts real (non-duplicate) units.
    """
    messages = []

    hu_by_container = {}
    for rec in all_hu_records:
        c = rec["container_no"]
        hu_by_container.setdefault(c, []).append(rec)

    for container, hu_list in hu_by_container.items():
        mask = df_orders["container_no"] == container
        if not mask.any():
            messages.append(f"[!] Container {container}: not found in Order List Excel")
            continue

        expected = int(df_orders.loc[mask, "units_expected"].sum())
        total = len(hu_list)
        # Duplicate HUs have pallet=0 AND pieces=0 (e.g. "0 FORGED BAR ON SKID")
        duplicates = sum(
            1 for r in hu_list
            if r.get("pallet", 1) == 0 and r.get("pieces", 1) == 0
        )
        actual = total - duplicates

        if expected == actual:
            if duplicates:
                messages.append(
                    f"[OK] {container}: {actual}/{expected} HUs "
                    f"(+ {duplicates} duplicate(s)) - PASS"
                )
            else:
                messages.append(f"[OK] {container}: {actual}/{expected} HUs - PASS")
        else:
            messages.append(
                f"[X] {container}: extracted {actual} HUs"
                + (f" (+ {duplicates} duplicate(s))" if duplicates else "")
                + f", expected {expected} - MISMATCH"
            )

    excel_containers = set(df_orders["container_no"].unique())
    parsed_containers = set(hu_by_container.keys())
    missing = excel_containers - parsed_containers
    for m in missing:
        exp = int(df_orders.loc[df_orders["container_no"] == m, "units_expected"].sum())
        messages.append(f"[!] {m}: no packing list uploaded ({exp} HUs expected)")

    return messages


# ═══════════════════════════════════════════════════════════════
# 6. ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════

def _detect_scanned_pages(pdf_path: str) -> tuple:
    """
    Open the PDF with pdfplumber and return (total_pages, scanned_indices).

    scanned_indices: 0-based indices of pages with zero extractable words
    (i.e. pure image / scanned pages).
    """
    scanned = []
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            words = page.extract_words()
            if not words:
                scanned.append(i)
    return total, scanned


from concurrent.futures import ThreadPoolExecutor

def process_shipment(
    order_path: str,
    arrival_paths: list,
    packing_paths: list,
    output_path: str,
) -> dict:
    """
    Main entry point. Processes all files and writes output Excel.

    Flow for each packing list PDF:
      1. Detect which pages have 0 extractable words  → scanned pages
      2. Run standard parse_packing_list()            → digital HU records
      2b. If filename had no container number, resolve container_no per HU
          from the Order List (delivery_no → container_no mapping).
      3. If scanned pages exist → extract_scanned_pages() via Gemini vision
      4. link_scanned_to_deliveries() assigns delivery_no using Order List
      5. Merge both sets into all_hu_records

    Returns summary dict with validation messages and row counts.
    """
    df_orders = parse_order_list(order_path)
    
    arrival = {}
    if arrival_paths:
        with ThreadPoolExecutor(max_workers=min(len(arrival_paths), 10)) as executor:
            results = list(executor.map(parse_arrival_notice_llm, arrival_paths))
            for res in results:
                arrival.update(res)

    all_hu_records = []
    parse_results = []

    # Build delivery_no → container_no lookup from the Order List once
    # (used in Step 2b when the packing list filename has no container number)
    delivery_to_container_ol = {
        str(row["receipt_po"]).strip(): str(row.get("container_no", "")).strip()
        for _, row in df_orders.iterrows()
        if str(row.get("container_no", "")).strip()
    }

    for pp in packing_paths:
        container = _derive_container_from_filename(pp)
        print(f"\n[*] Processing: {os.path.basename(pp)} (container: {container or '(none in filename)'})")

        # ── Step 1: Detect scanned (image-only) pages ──
        total_pages, scanned_indices = _detect_scanned_pages(pp)
        n_scanned = len(scanned_indices)
        n_digital = total_pages - n_scanned
        print(f"  Pages: {total_pages} total, {n_digital} digital, {n_scanned} scanned")

        # ── Step 2: Parse digital pages (standard extraction) ──
        hu_records, val_info = parse_packing_list(pp)
        print(f"  🗂  Digital extraction: {len(hu_records)} HU record(s)")

        # ── Step 2b: Resolve container_no when filename has none ──
        # parse_packing_list() sets container_no="" for every HU when the
        # filename contains no ISO container code (e.g. "Packing_lists.pdf").
        # Patch each record by looking up its delivery_no in the Order List.
        if not container and hu_records:
            patched = 0
            for rec in hu_records:
                if not rec.get("container_no"):
                    resolved = delivery_to_container_ol.get(rec.get("delivery_no", ""), "")
                    if resolved:
                        rec["container_no"] = resolved
                        patched += 1
            if patched:
                # Re-derive the dominant container from patched records
                containers_found = set(
                    r["container_no"] for r in hu_records if r.get("container_no")
                )
                container = next(iter(containers_found)) if len(containers_found) == 1 else ""
                print(
                    f"  [*] container_no resolved from Order List for {patched} HU(s): "
                    f"{sorted(containers_found)}"
                )
            else:
                print("  [!] Could not resolve container_no — delivery_nos not found in Order List")

        # ── Step 3: Extract scanned pages via Gemini vision ──
        scanned_records = []
        if scanned_indices:
            scanned_records = extract_scanned_pages(
                pdf_path=pp,
                scanned_indices=scanned_indices,
                container_no=container,
            )

            # ── Step 4: Link scanned HUs to deliveries via Order List ──
            if scanned_records:
                scanned_records = link_scanned_to_deliveries(
                    scanned_records=scanned_records,
                    df_orders=df_orders,
                    digital_records=hu_records,
                    container_no=container,
                )

        # ── Step 5: Merge digital + scanned ──
        combined = hu_records + scanned_records
        all_hu_records.extend(combined)

        parse_results.append({
            "file": os.path.basename(pp),
            "container": container,
            "hu_count": len(combined),
            "digital_hu_count": len(hu_records),
            "scanned_hu_count": len(scanned_records),
            "scanned_pages": n_scanned,
        })

    validation_messages = validate(df_orders, all_hu_records)
    df_final = merge_data(df_orders, arrival, all_hu_records)

    if not df_final.empty:
        for col in df_final.columns:
            df_final[col] = df_final[col].apply(
                lambda v: "" if pd.isna(v)
                else str(int(v)) if isinstance(v, float) and v == int(v)
                else str(v)
            )
        write_output_excel(df_final, output_path)

    return {
        "total_rows": len(df_final),
        "containers_parsed": len(parse_results),
        "parse_results": parse_results,
        "validation": validation_messages,
        "arrival_containers": list(arrival.keys()),
    }