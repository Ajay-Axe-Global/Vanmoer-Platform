# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# def write_excel(rows: list[dict], column_config: list[dict], output_path: str, formula_columns: dict | None = None):
#     """
#     Generic Excel writer shared by every client task.

#     Args:
#         rows:          List of dicts, one per output row. Keys match column_config field_key.
#         column_config: List of dicts defining columns: { "header": str, "field_key": str, "width": int, "num_format": str? }
#         output_path:   Where to save the .xlsx file.
#         formula_columns: Optional dict of field_key -> format string (e.g. "=B{row}*2") for computed cells.
#     """
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Output"

#     b        = Side(style="thin", color="BFBFBF")
#     border   = Border(left=b, right=b, top=b, bottom=b)
#     hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
#     hdr_fill = PatternFill("solid", start_color="1F4E79")
#     dat_font = Font(name="Arial", size=10)
#     even_fill= PatternFill("solid", start_color="D6E4F0")
#     odd_fill = PatternFill("solid", start_color="FFFFFF")
#     c_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     l_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

#     for ci, col in enumerate(column_config, 1):
#         cell = ws.cell(row=1, column=ci, value=col["header"])
#         cell.font = hdr_font; cell.fill = hdr_fill
#         cell.border = border; cell.alignment = c_align
#         ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col.get("width", 18)
#     ws.row_dimensions[1].height = 30

#     fcols = formula_columns or {}
#     for ri, row in enumerate(rows, 2):
#         fill = even_fill if ri % 2 == 0 else odd_fill
#         for ci, col in enumerate(column_config, 1):
#             fkey = col["field_key"]
#             val  = fcols[fkey].format(row=ri) if fkey in fcols else row.get(fkey, "")
#             cell = ws.cell(row=ri, column=ci, value=val)
#             cell.font = dat_font; cell.fill = fill
#             cell.border = border; cell.alignment = l_align
#             if "num_format" in col:
#                 cell.number_format = col["num_format"]
#         ws.row_dimensions[ri].height = 20

#     ws.freeze_panes = "A2"
#     wb.save(output_path)


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


def write_excel(rows: list[dict], column_config: list[dict], output_path: str, formula_columns: dict | None = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output"

    b        = Side(style="thin", color="BFBFBF")
    border   = Border(left=b, right=b, top=b, bottom=b)
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    dat_font = Font(name="Arial", size=10)
    even_fill= PatternFill("solid", start_color="D6E4F0")
    odd_fill = PatternFill("solid", start_color="FFFFFF")
    c_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_align  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for ci, col in enumerate(column_config, 1):
        cell = ws.cell(row=1, column=ci, value=col["header"])
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.border = border; cell.alignment = c_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col.get("width", 18)
    ws.row_dimensions[1].height = 30

    fcols = formula_columns or {}
    for ri, row in enumerate(rows, 2):
        fill = even_fill if ri % 2 == 0 else odd_fill
        for ci, col in enumerate(column_config, 1):
            fkey = col["field_key"]
            val  = fcols[fkey].format(row=ri) if fkey in fcols else row.get(fkey, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = dat_font; cell.fill = fill
            cell.border = border; cell.alignment = l_align
            if "num_format" in col:
                cell.number_format = col["num_format"]
        ws.row_dimensions[ri].height = 20

    # ── NEW: Create a proper Excel Table so column names are metadata ──
    if rows:
        last_col_letter = openpyxl.utils.get_column_letter(len(column_config))
        last_row = len(rows) + 1  # +1 for header row
        table_ref = f"A1:{last_col_letter}{last_row}"
        table = Table(displayName="DataTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    ws.freeze_panes = "A2"
    wb.save(output_path)